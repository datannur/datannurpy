"""Tests for Excel dataset validation and scanning."""

from __future__ import annotations

import codecs
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET
from zipfile import ZIP_DEFLATED, ZipFile

import pandas as pd
import pytest

from datannurpy import Catalog
from datannurpy.scanner.excel import _read_preview_rows, is_valid_tabular_dataset


def _write_xlsx(path: Path, rows: list[list[object]]) -> None:
    """Write rows to an xlsx file using openpyxl."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row in rows:
        ws.append(row)
    wb.save(path)


def _write_shared_strings_xlsx(path: Path) -> None:
    """Write a minimal xlsx using sharedStrings.xml."""
    files = {
        "xl/workbook.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Data" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>"
        ),
        "xl/_rels/workbook.xml.rels": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
            'Target="worksheets/sheet1.xml"/></Relationships>'
        ),
        "xl/worksheets/sheet1.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData><row r="1"><c r="A1" t="s"><v>0</v></c>'
            '<c r="B1" t="s"><v>2</v></c></row><row r="2">'
            '<c r="A2"><v>1</v></c><c r="B2" t="inlineStr">'
            "<is><t>Alice</t></is></c></row></sheetData></worksheet>"
        ),
        "xl/sharedStrings.xml": (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'count="3" uniqueCount="3"><si><t>id</t></si><si><t>unused</t></si>'
            "<si><t>name</t></si></sst>"
        ),
    }
    with ZipFile(path, "w", ZIP_DEFLATED) as zf:
        for name, content in files.items():
            zf.writestr(name, content)


def _write_minimal_xlsx(
    path: Path,
    *,
    workbook: str,
    rels: str,
    worksheet: str,
    shared_strings: str | None = None,
) -> None:
    """Write minimal xlsx package parts for parser edge-case tests."""
    with ZipFile(path, "w", ZIP_DEFLATED) as zf:
        zf.writestr("xl/workbook.xml", workbook)
        zf.writestr("xl/_rels/workbook.xml.rels", rels)
        zf.writestr("xl/worksheets/sheet1.xml", worksheet)
        if shared_strings is not None:
            zf.writestr("xl/sharedStrings.xml", shared_strings)


def _minimal_workbook(sheet_body: str) -> str:
    """Build a minimal workbook XML document."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheet_body}</sheets></workbook>"
    )


def _minimal_rels(target: str = "worksheets/sheet1.xml") -> str:
    """Build minimal workbook relationships XML."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'<Relationship Id="rId1" Target="{target}"/></Relationships>'
    )


def _minimal_worksheet(sheet_data: str) -> str:
    """Build a minimal worksheet XML document."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{sheet_data}</sheetData></worksheet>"
    )


class TestIsValidExcelDataset:
    """Tests for is_valid_tabular_dataset()."""

    def test_valid_dataset(self):
        rows = [("id", "name", "age"), (1, "Alice", 30), (2, "Bob", 25)]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid
        assert reason == ""

    def test_empty_rows(self):
        valid, reason = is_valid_tabular_dataset([])
        assert not valid
        assert reason == "empty sheet"

    def test_empty_header(self):
        valid, reason = is_valid_tabular_dataset([()])
        assert not valid
        assert reason == "empty header row"

    def test_does_not_start_at_a1(self):
        rows = [(None, None, "Code", "Name")]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "header does not start at column A"

    def test_none_gap_in_header(self):
        rows = [("Code", None, "Name", "Salary")]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "empty cells in header row"

    def test_duplicate_column_names(self):
        rows = [("Total", "Count", "Total")]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "duplicate column names"

    def test_non_text_header_numbers(self):
        rows = [(2023, 2024, 2025), (100, 200, 150)]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "non-text values in header row"

    def test_non_text_header_mixed(self):
        rows = [("Code", 42, "Name")]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "non-text values in header row"

    def test_data_wider_than_header(self):
        rows = [
            ("Rapport",),
            ("Code", "Nom", "Dept", "Salaire"),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "data wider than header row"

    def test_data_same_width_ok(self):
        rows = [
            ("id", "name", "salary"),
            (1, "Alice", 5000),
            (2, "Bob", 4500),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid

    def test_data_narrower_ok(self):
        rows = [
            ("id", "name", "salary"),
            (1, "Alice", None),
            (2, None, None),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid

    def test_data_row_all_none_ignored(self):
        rows = [
            ("id", "name"),
            (None, None),
            (1, "Alice"),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid

    def test_empty_tuple_row_ignored(self):
        rows = [
            ("id", "name"),
            (),
            (1, "Alice"),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid

    def test_title_then_wider_data(self):
        """Title in row 1, real header wider → detected by criterion 5."""
        rows = [
            ("Rapport 2024", None, None),
            ("Code", "Nom", "Salaire"),
            (101, "Ventes", 5000),
        ]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid
        assert reason == "data wider than header row"

    def test_merged_cells_produce_none(self):
        """Merged cells in openpyxl read-only produce None → detected."""
        rows = [("Total", None, "Total", None)]
        valid, reason = is_valid_tabular_dataset(rows)
        assert not valid

    def test_single_column_dataset(self):
        rows = [("id",), (1,), (2,)]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid

    def test_header_only_no_data(self):
        rows = [("id", "name", "salary")]
        valid, reason = is_valid_tabular_dataset(rows)
        assert valid


class TestScanExcelValidation:
    """Integration tests: invalid Excel files are skipped in scan."""

    def test_read_excel_detects_html_xls_with_bom(self, tmp_path: Path, capsys):
        """HTML renamed to .xls should still be detected after a BOM."""
        from datannurpy.scanner.excel import read_excel

        xls_path = tmp_path / "report.xls"
        xls_path.write_bytes(
            codecs.BOM_UTF8 + b"\n\n<!DOCTYPE html><html><body>report</body></html>"
        )

        result = read_excel(xls_path, quiet=False)

        captured = capsys.readouterr()
        assert result is None
        assert "HTML content detected" in captured.err

    def test_xlsx_with_title_row_skipped(self, tmp_path: Path, capsys):
        """xlsx with a title row should be skipped."""
        _write_xlsx(
            tmp_path / "report.xlsx",
            [
                ["Rapport annuel 2024", None, None],
                ["Code", "Nom", "Salaire"],
                [101, "Alice", 5000],
            ],
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "not a valid tabular dataset" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_xlsx_with_numeric_header_skipped(self, tmp_path: Path, capsys):
        """xlsx with numbers in header should be skipped."""
        _write_xlsx(
            tmp_path / "pivot.xlsx",
            [
                [2023, 2024, 2025],
                [100, 200, 150],
            ],
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "not a valid tabular dataset" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_xlsx_with_duplicate_headers_skipped(self, tmp_path: Path, capsys):
        """xlsx with duplicate column names should be skipped."""
        _write_xlsx(
            tmp_path / "dupes.xlsx",
            [
                ["Total", "Count", "Total"],
                [100, 5, 200],
            ],
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "not a valid tabular dataset" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_xlsx_valid_dataset_scanned(self, tmp_path: Path):
        """Valid xlsx should be scanned normally."""
        df = pd.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"], "age": [30, 25]})
        df.to_excel(tmp_path / "valid.xlsx", index=False)

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=True)

        assert len(catalog.dataset.all()) == 1
        assert len(catalog.variable.all()) == 3

    def test_xlsx_data_wider_than_header_skipped(self, tmp_path: Path, capsys):
        """xlsx where data extends beyond header width should be skipped."""
        _write_xlsx(
            tmp_path / "wider.xlsx",
            [
                ["Titre"],
                ["Code", "Nom", "Salaire"],
                [101, "Alice", 5000],
            ],
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "not a valid tabular dataset" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_schema_mode_xlsx_invalid_skipped(self, tmp_path: Path):
        """Invalid xlsx should return no variables in schema mode."""
        _write_xlsx(
            tmp_path / "pivot.xlsx",
            [
                [2023, 2024, 2025],
                [100, 200, 150],
            ],
        )

        catalog = Catalog(depth="variable")
        catalog.add_folder(tmp_path, quiet=True)

        assert len(catalog.variable.all()) == 0

    def test_schema_mode_xlsx_valid_scanned(self, tmp_path: Path):
        """Valid xlsx should return variables in schema mode."""
        df = pd.DataFrame({"id": [1, 2], "name": ["Alice", "Bob"]})
        df.to_excel(tmp_path / "valid.xlsx", index=False)

        catalog = Catalog(depth="variable")
        catalog.add_folder(tmp_path, quiet=True)

        assert len(catalog.variable.all()) == 2

    def test_schema_mode_xlsx_shared_strings_fast_path(self, tmp_path: Path):
        """Schema-only xlsx preview should resolve shared strings locally."""
        _write_shared_strings_xlsx(tmp_path / "shared.xlsx")

        catalog = Catalog(depth="variable")
        catalog.add_folder(tmp_path, quiet=True)

        assert [v.name for v in catalog.variable.all()] == ["id", "name"]

    def test_xlsx_preview_falls_back_to_openpyxl(self, tmp_path: Path, monkeypatch):
        """Fast preview failures should fall back to openpyxl."""
        xlsx_path = tmp_path / "valid.xlsx"
        _write_xlsx(xlsx_path, [["id", "name"], [1, "Alice"]])

        from datannurpy.scanner import excel as excel_mod

        def fail_fast(_path: Path) -> list[tuple[object, ...]]:
            raise ValueError("boom")

        monkeypatch.setattr(excel_mod, "_read_xlsx_preview_rows_fast", fail_fast)

        rows = _read_preview_rows(xlsx_path, quiet=True)

        assert rows == [("id", "name"), (1, "Alice")]

    def test_xlsx_preview_non_path_source_uses_openpyxl(self):
        """Non-path xlsx sources should keep using openpyxl directly."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        rows = _read_preview_rows(buffer, quiet=True)

        assert rows == [("id", "name"), (1, "Alice")]

    def test_xlsx_fast_preview_sparse_and_scalar_values(self, tmp_path: Path):
        """Fast preview should preserve sparse rows and scalar cell types."""
        from datannurpy.scanner.excel import _read_xlsx_preview_rows_fast

        xlsx_path = tmp_path / "sparse.xlsx"
        _write_minimal_xlsx(
            xlsx_path,
            workbook=_minimal_workbook('<sheet name="Data" sheetId="1" r:id="rId1"/>'),
            rels=_minimal_rels("/xl/worksheets/sheet1.xml"),
            worksheet=_minimal_worksheet(
                '<row r="2"><ext/><c r="A2" t="str"><v>id</v></c>'
                '<c r="B2" t="b"><v>1</v></c><c r="C2"/>'
                '<c r="D2"><v>3.5</v></c></row>'
                '<row r="11"><c r="A11"><v>99</v></c></row>'
            ),
        )

        assert _read_xlsx_preview_rows_fast(xlsx_path) == [(), ("id", True, None, 3.5)]

    def test_xlsx_fast_preview_empty_sheet(self, tmp_path: Path):
        """Fast preview should return no rows for an empty worksheet."""
        from datannurpy.scanner.excel import _read_xlsx_preview_rows_fast

        xlsx_path = tmp_path / "empty.xlsx"
        _write_minimal_xlsx(
            xlsx_path,
            workbook=_minimal_workbook('<sheet name="Data" sheetId="1" r:id="rId1"/>'),
            rels=_minimal_rels(),
            worksheet=_minimal_worksheet(""),
        )

        assert _read_xlsx_preview_rows_fast(xlsx_path) == []

    def test_xlsx_fast_preview_invalid_workbook_fails(self, tmp_path: Path):
        """Malformed workbook relationships should fail fast for fallback."""
        from datannurpy.scanner.excel import _read_xlsx_preview_rows_fast

        rels = _minimal_rels()
        worksheet = _minimal_worksheet("")

        no_sheets = tmp_path / "no_sheets.xlsx"
        _write_minimal_xlsx(
            no_sheets,
            workbook=_minimal_workbook(""),
            rels=rels,
            worksheet=worksheet,
        )
        with pytest.raises(ValueError, match="no sheets"):
            _read_xlsx_preview_rows_fast(no_sheets)

        no_rel_id = tmp_path / "no_rel_id.xlsx"
        _write_minimal_xlsx(
            no_rel_id,
            workbook=_minimal_workbook('<sheet name="Data" sheetId="1"/>'),
            rels=rels,
            worksheet=worksheet,
        )
        with pytest.raises(ValueError, match="no relationship id"):
            _read_xlsx_preview_rows_fast(no_rel_id)

        missing_rel = tmp_path / "missing_rel.xlsx"
        _write_minimal_xlsx(
            missing_rel,
            workbook=_minimal_workbook('<sheet name="Data" sheetId="1" r:id="rId2"/>'),
            rels=rels,
            worksheet=worksheet,
        )
        with pytest.raises(ValueError, match="relationship is missing"):
            _read_xlsx_preview_rows_fast(missing_rel)

    def test_xlsx_fast_preview_incomplete_shared_strings(self, tmp_path: Path):
        """Incomplete sharedStrings should fail so openpyxl can handle fallback."""
        from datannurpy.scanner.excel import _read_xlsx_preview_rows_fast

        xlsx_path = tmp_path / "bad_shared.xlsx"
        _write_minimal_xlsx(
            xlsx_path,
            workbook=_minimal_workbook('<sheet name="Data" sheetId="1" r:id="rId1"/>'),
            rels=_minimal_rels(),
            worksheet=_minimal_worksheet(
                '<row r="1"><c r="A1" t="s"><v>1</v></c></row>'
            ),
            shared_strings=(
                '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
                "<si><t>id</t></si></sst>"
            ),
        )

        with pytest.raises(ValueError, match="incomplete"):
            _read_xlsx_preview_rows_fast(xlsx_path)

    def test_xlsx_parser_scalar_helpers(self):
        """Small scalar helpers should cover non-file parser branches."""
        from datannurpy.scanner import excel as excel_mod

        inline_empty = ET.fromstring('<c t="inlineStr"><v>ignored</v></c>')
        shared_indices: set[int] = set()

        assert excel_mod._column_index(None, 7) == 7
        assert excel_mod._column_index("AB", 7) == 28
        assert excel_mod._coerce_xlsx_number("not-a-number") == "not-a-number"
        assert excel_mod._xlsx_child_text(ET.fromstring("<c><x /></c>"), "v") is None
        assert excel_mod._read_xlsx_cell_value(ET.fromstring("<c />"), set()) is None
        assert excel_mod._read_xlsx_cell_value(inline_empty, shared_indices) == ""
        assert (
            excel_mod._read_xlsx_cell_value(
                ET.fromstring('<c t="e"><v>#N/A</v></c>'), set()
            )
            == "#N/A"
        )

    def test_xls_empty_file(self, tmp_path: Path):
        """Empty .xls file should produce no variables."""
        (tmp_path / "empty.xls").write_bytes(b"")

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=True)

        assert len(catalog.variable.all()) == 0

    def test_xls_corrupted_file(self, tmp_path: Path, capsys):
        """Corrupted .xls file should warn and produce no variables."""
        (tmp_path / "bad.xls").write_bytes(b"not a real xls file")

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "bad.xls" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_xls_html_report_skipped_with_clear_message(self, tmp_path: Path, capsys):
        """HTML content renamed to .xls should be classified as untreatable."""
        (tmp_path / "report.xls").write_bytes(
            b"<!DOCTYPE html><html><body>report</body></html>"
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "HTML content detected" in captured.err
        assert "skipped as untreatable" in captured.err
        assert len(catalog.variable.all()) == 0

    def test_nested_xls_html_warning_uses_relative_path(self, tmp_path: Path, capsys):
        """HTML .xls warnings from add_folder should keep the relative path."""
        nested = tmp_path / "folder" / "subfolder"
        nested.mkdir(parents=True)
        (nested / "report.xls").write_bytes(
            b"<!DOCTYPE html><html><body>report</body></html>"
        )

        catalog = Catalog()
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "folder/subfolder/report.xls" in captured.err
        assert "⚠ report.xls:" not in captured.err

    def test_xls_invalid_header_skipped(self, tmp_path: Path, monkeypatch):
        """xls with numeric columns should be skipped via post-read validation."""
        from datannurpy.scanner import excel as excel_mod

        xls_path = tmp_path / "pivot.xls"
        xls_path.write_bytes(b"dummy")

        numeric_df = pd.DataFrame({2023: [100], 2024: [200]})
        monkeypatch.setattr(excel_mod, "read_excel", lambda *_a, **_kw: numeric_df)

        vars_, count, frequency = excel_mod.scan_excel(
            xls_path, dataset_id="test---pivot_xls"
        )
        assert vars_ == []
        assert count is None
        assert frequency is None

    def test_xls_empty_sheet(self, tmp_path: Path, monkeypatch):
        """xls with empty sheet should return None from read_excel."""
        from datannurpy.scanner.excel import read_excel

        xls_path = tmp_path / "empty_sheet.xls"
        xls_path.write_bytes(b"x" * 10)

        monkeypatch.setattr(pd, "read_excel", lambda *_a, **_kw: pd.DataFrame())

        result = read_excel(xls_path)
        assert result is None

    def test_read_excel_captures_openpyxl_warning_as_debug(
        self, tmp_path: Path, monkeypatch, capsys
    ):
        """openpyxl parser warnings should not leak as raw terminal output."""
        import warnings

        from datannurpy.scanner.excel import read_excel
        from datannurpy.utils import configure_logging

        xlsx_path = tmp_path / "style.xlsx"
        xlsx_path.write_bytes(b"dummy")
        log_path = tmp_path / "scan.log"

        def fake_read_excel(*_args, **_kwargs):
            warnings.warn(
                "Workbook contains no default style, apply openpyxl's default",
                UserWarning,
                stacklevel=2,
            )
            return pd.DataFrame({"id": [1]})

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)

        configure_logging(verbose=False, log_file=log_path)
        try:
            result = read_excel(xlsx_path, quiet=False, path_label="folder/style.xlsx")
        finally:
            configure_logging()

        captured = capsys.readouterr()
        log_text = log_path.read_text()
        assert result is not None
        assert "Workbook contains no default style" not in captured.err
        assert "folder/style.xlsx: Excel parser diagnostic" in log_text
        assert "Workbook contains no default style" in log_text

    def test_read_excel_captures_xlrd_stdout_as_debug(
        self, tmp_path: Path, monkeypatch, capsys
    ):
        """xlrd stdout diagnostics should not leak as raw terminal output."""
        from datannurpy.scanner.excel import read_excel
        from datannurpy.utils import configure_logging

        xls_path = tmp_path / "ole2.xls"
        xls_path.write_bytes(b"not html")
        log_path = tmp_path / "scan.log"

        def fake_read_excel(*_args, **_kwargs):
            print(
                "WARNING *** OLE2 inconsistency: SSCS size is 0 but SSAT size is non-zero"
            )
            return pd.DataFrame({"id": [1]})

        monkeypatch.setattr(pd, "read_excel", fake_read_excel)

        configure_logging(verbose=False, log_file=log_path)
        try:
            result = read_excel(xls_path, quiet=False, path_label="folder/ole2.xls")
        finally:
            configure_logging()

        captured = capsys.readouterr()
        log_text = log_path.read_text()
        assert result is not None
        assert captured.out == ""
        assert "OLE2 inconsistency" not in captured.err
        assert "folder/ole2.xls: Excel parser diagnostic" in log_text
        assert "OLE2 inconsistency" in log_text

    def test_schema_mode_xls_invalid_skipped(self, tmp_path: Path, monkeypatch):
        """Schema-only scan of .xls with invalid header returns no variables."""
        xls_path = tmp_path / "pivot.xls"
        xls_path.write_bytes(b"dummy")

        numeric_df = pd.DataFrame({2023: [100], 2024: [200]})
        monkeypatch.setattr(pd, "read_excel", lambda *_a, **_kw: numeric_df)

        catalog = Catalog(depth="variable")
        catalog.add_folder(tmp_path, quiet=True)

        assert len(catalog.variable.all()) == 0

    def test_schema_mode_xls_html_report_skipped(self, tmp_path: Path, capsys):
        """Schema-only scan should skip HTML reports renamed to .xls."""
        (tmp_path / "report.xls").write_bytes(b"<html><body>report</body></html>")

        catalog = Catalog(depth="variable")
        catalog.add_folder(tmp_path, quiet=False)

        captured = capsys.readouterr()
        assert "HTML content detected" in captured.err
        assert len(catalog.variable.all()) == 0


class TestScanExcelSchemaRemote:
    """Test remote schema-only Excel validation."""

    def test_remote_xlsx_invalid_skipped(self, tmp_path: Path):
        """Remote invalid xlsx should copy locally before schema scanning."""
        from unittest.mock import MagicMock

        import openpyxl

        from datannurpy.scanner.scan import scan_file

        local_path = tmp_path / "pivot.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.append([2023, 2024, 2025])
        ws.append([100, 200, 150])
        wb.save(local_path)

        mock_fs = MagicMock()
        mock_fs.is_local = False
        mock_fs.ensure_local.return_value.__enter__ = MagicMock(return_value=local_path)
        mock_fs.ensure_local.return_value.__exit__ = MagicMock(return_value=None)

        result = scan_file(
            Path("/remote/pivot.xlsx"),
            "excel",
            dataset_id="test---pivot_xlsx",
            schema_only=True,
            fs=mock_fs,
        )

        assert len(result.variables) == 0
        assert result.nb_row is None
        # Downloaded under a safe temp name carrying the resolved format's extension.
        mock_fs.ensure_local.assert_called_once_with(
            str(Path("/remote/pivot.xlsx")), "data.xlsx"
        )

    def test_remote_xls_html_schema_skipped_without_download(self, capsys):
        """Remote HTML renamed to .xls should be skipped before full download."""
        from io import BytesIO
        from unittest.mock import MagicMock

        from datannurpy.scanner.scan import scan_file

        mock_fs = MagicMock()
        mock_fs.is_local = False
        mock_fs.open.return_value.__enter__ = MagicMock(
            return_value=BytesIO(b"<!DOCTYPE html><html><body>report</body></html>")
        )
        mock_fs.open.return_value.__exit__ = MagicMock(return_value=None)

        result = scan_file(
            Path("/remote/report.xls"),
            "excel",
            dataset_id="test---report_xls",
            schema_only=True,
            fs=mock_fs,
        )

        captured = capsys.readouterr()
        assert result.variables == []
        assert result.nb_row is None
        mock_fs.ensure_local.assert_not_called()
        assert "HTML content detected" in captured.err

    def test_remote_xls_html_full_scan_skipped_without_download(self, capsys):
        """Remote HTML renamed to .xls should be skipped before full download."""
        from io import BytesIO
        from unittest.mock import MagicMock

        from datannurpy.scanner.scan import scan_file

        mock_fs = MagicMock()
        mock_fs.is_local = False
        mock_fs.open.return_value.__enter__ = MagicMock(
            return_value=BytesIO(b"<html><body>report</body></html>")
        )
        mock_fs.open.return_value.__exit__ = MagicMock(return_value=None)

        result = scan_file(
            Path("/remote/report.xls"),
            "excel",
            dataset_id="test---report_xls",
            fs=mock_fs,
            quiet=False,
        )

        captured = capsys.readouterr()
        assert result.variables == []
        assert result.nb_row is None
        mock_fs.ensure_local.assert_not_called()
        assert "HTML content detected" in captured.err

    def test_remote_xls_non_html_uses_download_path(self, monkeypatch):
        """Remote .xls files that are not HTML should continue to the local scan path."""
        from io import BytesIO
        from unittest.mock import MagicMock

        from datannurpy.scanner import scan as scan_mod

        mock_fs = MagicMock()
        mock_fs.is_local = False
        mock_fs.open.return_value.__enter__ = MagicMock(
            return_value=BytesIO(b"not html")
        )
        mock_fs.open.return_value.__exit__ = MagicMock(return_value=None)

        local_path = Path("/tmp/local.xls")
        ensure_local_ctx = MagicMock()
        ensure_local_ctx.__enter__.return_value = local_path
        ensure_local_ctx.__exit__.return_value = None
        mock_fs.ensure_local.return_value = ensure_local_ctx

        expected = scan_mod.ScanResult(variables=[], nb_row=0)
        monkeypatch.setattr(scan_mod, "_scan_local", lambda *args, **kwargs: expected)

        result = scan_mod.scan_file(
            Path("/remote/report.xls"),
            "excel",
            dataset_id="test---report_xls",
            fs=mock_fs,
        )

        assert result == expected
        mock_fs.ensure_local.assert_called_once()


class TestScanOds:
    """OpenDocument spreadsheets ride the Excel pipeline via the pandas odf engine."""

    @staticmethod
    def _write_ods(path: Path, df: pd.DataFrame) -> None:
        with pd.ExcelWriter(path, engine="odf") as writer:
            df.to_excel(writer, index=False)

    def test_value_depth(self, tmp_path: Path):
        ods = tmp_path / "sales.ods"
        self._write_ods(
            ods, pd.DataFrame({"city": ["Bern", "Sion"], "amount": [10, 20]})
        )
        catalog = Catalog(quiet=True)
        catalog.add_dataset(ods)
        ds = catalog.dataset.all()[0]
        assert ds.delivery_format == "ods"
        assert ds.nb_row == 2
        assert [v.name for v in catalog.variable.all()] == ["city", "amount"]

    def test_variable_depth_via_folder(self, tmp_path: Path):
        """Folder discovery picks .ods up; schema-only reads through pandas/odf."""
        self._write_ods(
            tmp_path / "sales.ods", pd.DataFrame({"city": ["Bern"], "amount": [10]})
        )
        catalog = Catalog(depth="variable", quiet=True)
        catalog.add_folder(tmp_path)
        ds = catalog.dataset.all()[0]
        assert ds.delivery_format == "ods"
        assert ds.nb_row is None  # not scanned at variable depth
        assert [v.name for v in catalog.variable.all()] == ["city", "amount"]

    def test_empty_sheet_scans_as_zero_rows(self, tmp_path: Path):
        """No streaming preflight for .ods: an empty sheet resolves post-read."""
        ods = tmp_path / "empty.ods"
        self._write_ods(ods, pd.DataFrame())
        catalog = Catalog(quiet=True)
        catalog.add_dataset(ods)
        assert catalog.dataset.all()[0].nb_row == 0

    def test_invalid_tabular_skipped(self, tmp_path: Path, capsys):
        """Numeric headers fail the post-read validation, like .xls."""
        ods = tmp_path / "pivot.ods"
        self._write_ods(ods, pd.DataFrame({2023: [100], 2024: [200]}))
        catalog = Catalog(quiet=True)
        catalog.add_dataset(ods, quiet=False)
        captured = capsys.readouterr()
        assert "not a valid tabular dataset" in captured.err
        assert len(catalog.variable.all()) == 0


class TestReadExcelHelper:
    """read_excel (used by metadata loading) returns None on empty or broken input."""

    def test_zero_byte_file_returns_none(self, tmp_path: Path):
        from datannurpy.scanner.excel import read_excel

        empty = tmp_path / "empty.xlsx"
        empty.write_bytes(b"")
        assert read_excel(empty, quiet=True) is None

    def test_read_error_returns_none(self, tmp_path: Path, capsys):
        from datannurpy.scanner.excel import read_excel

        bad = tmp_path / "bad.xlsx"
        bad.write_bytes(b"not a real excel file")
        assert read_excel(bad, quiet=False) is None
        assert "✗  bad.xlsx" in capsys.readouterr().err
